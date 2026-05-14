from openpyxl import load_workbook
import re
import html

wb_path = r"C:\Users\Carlos Costato\OneDrive - HITSS DO BRASIL SERVIÇOS TECNOLOGICOS LTDA\Documents\Formulário_noc\Form_novo.xlsx"
sheet_name = 'Solicitação de acesso'

wb = load_workbook(filename=wb_path, data_only=True)
if sheet_name not in wb.sheetnames:
    print(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
    raise SystemExit(1)

ws = wb[sheet_name]
# Read across columns B..BZ (2..78) for rows 2..200 and concatenate non-empty cells per row
labels = []
for r in range(2, 201):
    parts = []
    for c in range(2, 79):
        v = ws.cell(row=r, column=c).value
        if v is None: continue
        s = str(v).strip()
        if s == '': continue
        parts.append(s)
    if not parts:
        continue
    # join with space preserving existing line breaks in parts
    row_text = ' '.join([p.replace('\r',' ').replace('\n',' ') for p in parts]).strip()
    if row_text:
        labels.append(row_text)

# Basic grouping: new section when label starts with digit
sections = []
current = { 'title': 'Informações', 'fields': [] }
for lbl in labels:
    m = re.match(r"^(\d+(?:\.\d+)*)\s*-?\s*(.*)$", lbl)
    if m:
        # starts with number -> new section
        num = m.group(1)
        title = m.group(2) or lbl
        if current and (current.get('fields')):
            sections.append(current)
        current = { 'title': f"{num} {title}", 'fields': [] }
    else:
        # treat as field or paragraph if very long
        if len(lbl) > 120 or '\n' in lbl:
            # paragraph
            current['fields'].append({ 'type': 'textarea', 'label': lbl })
        else:
            # short label -> text input
            current['fields'].append({ 'type': 'text', 'label': lbl })

if current and current.get('fields'):
    sections.append(current)

# Build HTML using project's CSS pattern (concise)
html_parts = []
html_parts.append("<!DOCTYPE html>\n<html lang=\"pt-BR\">\n<head>\n<meta charset=\"utf-8\">\n<meta name=\"viewport\" content=\"width=device-width,initial-scale=1\">\n<title>Formulário de Acesso - Gerado</title>\n<style>\n*{margin:0;padding:0;box-sizing:border-box}body{font-family:'Segoe UI',Tahoma,Arial;background:linear-gradient(135deg,#667eea 0%,#764ba2 100%);padding:20px} .container{max-width:1000px;margin:0 auto;background:#fff;border-radius:15px;box-shadow:0 20px 60px rgba(0,0,0,.3)}.header{background:linear-gradient(135deg,#667eea 0%,#764ba2 100%);color:#fff;padding:40px;text-align:center}.form-content{padding:40px}.section{margin-bottom:35px;padding-bottom:25px;border-bottom:2px solid #f0f0f0}.section-title{color:#667eea;font-size:20px;font-weight:600;margin-bottom:20px;display:flex;align-items:center;gap:10px}.section-number{background:#667eea;color:#fff;width:30px;height:30px;border-radius:50%;display:flex;align-items:center;justify-content:center;font-weight:700}.form-group{margin-bottom:20px}label{display:block;margin-bottom:8px;color:#333;font-weight:600}input,textarea,select{width:100%;padding:12px;border:2px solid #e0e0e0;border-radius:8px} .submit-btn{background:linear-gradient(135deg,#667eea 0%,#764ba2 100%);color:#fff;padding:15px;border:none;border-radius:8px;font-weight:700;width:100%} </style>\n</head>\n<body>\n<div class=\"container\">\n<div class=\"header\">\n<h1 style=\"margin:0\">SOLICITAÇÃO DE ACESSO</h1>\n<p>Gerado a partir de Form_novo.xlsx — aba: Solicitação de acesso</p>\n</div>\n<div class=\"form-content\">\n<form id=\"acessoForm\">")

sec_idx = 1
for sec in sections:
    html_parts.append(f"<div class=\"section\">\n  <div class=\"section-title\"><span class=\"section-number\">{sec_idx}</span>{html.escape(sec['title'])}</div>")
    for field in sec['fields']:
        label = html.escape(field['label']).replace('\n','<br>')
        if field['type'] == 'textarea':
            html_parts.append(f"  <div class=\"form-group\">\n    <label>{label}</label>\n    <textarea rows=\"4\"></textarea>\n  </div>")
        else:
            # text input
            # generate input name from label
            name = re.sub(r"[^0-9a-zA-Z]+","_", field['label']).strip('_').lower()[:40]
            html_parts.append(f"  <div class=\"form-group\">\n    <label>{label}</label>\n    <input type=\"text\" name=\"{name}\" placeholder=\"\">\n  </div>")
    html_parts.append("</div>")
    sec_idx += 1

html_parts.append("<button type=\"submit\" class=\"submit-btn\">✓ Clique e Baixe PDF e XLSX</button>\n</form>\n</div>\n</div>\n</body>\n</html>")

out_path = r"C:\Users\Carlos Costato\OneDrive - HITSS DO BRASIL SERVIÇOS TECNOLOGICOS LTDA\Documents\Formulário_noc\Form_de_acesso.html"
with open(out_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(html_parts))

print(f"Wrote {out_path} with {len(sections)} sections and total fields {sum(len(s['fields']) for s in sections)}")
