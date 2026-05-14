import json
from openpyxl import load_workbook
import sys

wb_path = r"C:\Users\Carlos Costato\OneDrive - HITSS DO BRASIL SERVI\CIOS TECNOLOGICOS LTDA\Documents\Formulário_noc\Form_novo.xlsx"
# The above path seems broken due to line break in folder name; fallback to relative
wb_path = r"C:\Users\Carlos Costato\OneDrive - HITSS DO BRASIL SERVIÇOS TECNOLOGICOS LTDA\Documents\Formulário_noc\Form_novo.xlsx"

try:
    wb = load_workbook(filename=wb_path, data_only=True)
except Exception as e:
    print(json.dumps({"error": f"Cannot open workbook: {e}"}))
    sys.exit(1)

sheet_name = 'Solicitação de acesso'
if sheet_name not in wb.sheetnames:
    print(json.dumps({"error": f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"}))
    sys.exit(1)

ws = wb[sheet_name]
rows = []
for r in ws.iter_rows(min_row=2, max_row=200, min_col=2, max_col=78, values_only=True):
    rows.append([None if v is None else str(v) for v in r])

print(json.dumps({"sheet": sheet_name, "rows": rows}, ensure_ascii=False))
